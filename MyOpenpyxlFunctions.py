import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

# function to delete content of an Excel sheet keeping headers
def delete_sheet(workbookName, sheetName):
    wb = load_workbook(workbookName)
    sheet = wb[sheetName]
    while(sheet.max_row>1):
        sheet.delete_rows(2)
    return

# function to fully delete content of an Excel sheet
def delete_sheet_all(workbookName, sheetName):
    wb = load_workbook(workbookName)
    sheet = wb[sheetName]
    while(sheet.max_row>1):
        sheet.delete_rows(2)
    sheet.delete_row(1)
    return

# function to write a pandas df into an Excel sheet
def copy_df_in_excel_sheet(workbookName, sheetName, df):
    wb = load_workbook(workbookName)
    sheet = wb[sheetName]
    for row in dataframe_to_rows(df, index=False,header=False): #set header to True if you want to keep df headers
            sheet.append(row)
    return

